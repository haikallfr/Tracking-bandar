#!/usr/bin/env python3

from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "storage" / "app.sqlite"
SECTOR_DIR = ROOT / "assets" / "Sektor"


def sector_from_filename(name: str) -> str:
    match = re.search(r"Daftar Saham\s*-\s*(.+?)\s*-\s*\d{8}\.xlsx$", name)
    if not match:
        raise ValueError(f"Nama file tidak sesuai pola: {name}")
    return match.group(1).strip()


def normalize_subsector(sector: str, company: str) -> str:
    company_l = company.lower()
    sector_l = sector.lower()

    if sector_l == "financials":
        if "bank" in company_l:
            return "Banks"
        if "asuransi" in company_l or "insurance" in company_l:
            return "Insurance"
        if "sekuritas" in company_l or "securities" in company_l:
            return "Securities"
        if "multi finance" in company_l or "finance" in company_l or "pembiayaan" in company_l:
            return "Financing Service"
    if sector_l == "healthcare":
        if "laborator" in company_l:
            return "Healthcare Services"
        if "farm" in company_l or "kimia farma" in company_l or "kalbe" in company_l:
            return "Pharmaceuticals"
        if "hospital" in company_l or "medika" in company_l or "hermina" in company_l:
            return "Healthcare Providers"
    if sector_l == "energy":
        if "gas" in company_l:
            return "Gas"
        if "coal" in company_l or "bukit asam" in company_l:
            return "Coal"
        if "geothermal" in company_l:
            return "Geothermal"
    return sector


def ensure_symbol_rows(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS symbol_reference (
            symbol TEXT PRIMARY KEY,
            company_name TEXT NOT NULL DEFAULT "",
            listing_board TEXT NOT NULL DEFAULT "",
            listed_shares REAL NOT NULL DEFAULT 0,
            sector TEXT NOT NULL DEFAULT "",
            subsector TEXT NOT NULL DEFAULT "",
            source TEXT NOT NULL DEFAULT "",
            updated_at TEXT NOT NULL
        )
        """
    )


def import_workbooks() -> int:
    ensure = sqlite3.connect(DB_PATH)
    ensure_symbol_rows(ensure)
    ensure.close()

    conn = sqlite3.connect(DB_PATH)
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    total = 0

    files = sorted(SECTOR_DIR.glob("*.xlsx"))
    for path in files:
        sector = sector_from_filename(path.name)
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or len(row) < 6:
                continue

            symbol = str(row[1] or "").strip().upper()
            company = str(row[2] or "").strip()
            listing_board = str(row[5] or "").strip()
            shares_raw = str(row[4] or "0").replace(".", "").replace(",", ".").strip()
            try:
                listed_shares = float(shares_raw)
            except ValueError:
                listed_shares = 0.0

            if not symbol:
                continue

            subsector = normalize_subsector(sector, company)

            conn.execute(
                """
                INSERT INTO symbol_reference(symbol, company_name, listing_board, listed_shares, sector, subsector, source, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(symbol) DO UPDATE SET
                    company_name = excluded.company_name,
                    listing_board = excluded.listing_board,
                    listed_shares = excluded.listed_shares,
                    sector = excluded.sector,
                    subsector = excluded.subsector,
                    source = excluded.source,
                    updated_at = excluded.updated_at
                """,
                (
                    symbol,
                    company,
                    listing_board,
                    listed_shares,
                    sector,
                    subsector,
                    "assets_sector_xlsx",
                    now,
                ),
            )
            total += 1

    conn.commit()
    conn.close()
    return total


if __name__ == "__main__":
    imported = import_workbooks()
    print(f"Imported {imported} sector rows from assets/Sektor.")
